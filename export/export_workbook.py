#!/usr/bin/env python3
"""Export SQLite database to formatted Excel workbook."""

import sqlite3
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE, "db", "curriculum.db")
OUT_PATH = os.path.join(BASE, "export", "alignment_report.xlsx")

GREEN = PatternFill('solid', fgColor='DCFCE7')
RED = PatternFill('solid', fgColor='FEE2E2')
YELLOW = PatternFill('solid', fgColor='FEF9C3')
BLUE_LIGHT = PatternFill('solid', fgColor='DBEAFE')
GRAY = PatternFill('solid', fgColor='F3F4F6')
HEADER_FILL = PatternFill('solid', fgColor='1E40AF')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
BOLD = Font(bold=True, size=11)
MONO = Font(name='Consolas', size=10)
SMALL = Font(size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='DEE2E6'),
    right=Side(style='thin', color='DEE2E6'),
    top=Side(style='thin', color='DEE2E6'),
    bottom=Side(style='thin', color='DEE2E6')
)

def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def auto_width(ws, min_w=8, max_w=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

def export():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "1E40AF"

    ws['A1'] = "RMS Standards Alignment Report"
    ws['A1'].font = Font(bold=True, size=16, color='1E40AF')
    ws['A2'] = "Drummond Math Solutions · CPM Curriculum Coverage Analysis"
    ws['A2'].font = Font(size=11, color='6B7280')
    ws['A4'] = "Coverage by Course & Framework"
    ws['A4'].font = BOLD

    frameworks = ['MN-2022', 'CCSS-M', 'MN-2007', 'TEKS']
    courses = [r['id'] for r in conn.execute("SELECT id FROM cpm_courses ORDER BY id").fetchall()]

    row = 6
    headers = ['Course', 'Modules'] + [f'{fw} Covered' for fw in frameworks] + [f'{fw} Gaps' for fw in frameworks]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))

    for course_id in courses:
        row += 1
        total_mods = conn.execute("SELECT COUNT(*) FROM cpm_modules WHERE course_id=?", (course_id,)).fetchone()[0]
        ws.cell(row=row, column=1, value=course_id).font = BOLD
        ws.cell(row=row, column=2, value=total_mods)

        col = 3
        for fw in frameworks:
            total_stds = conn.execute("SELECT COUNT(*) FROM standards WHERE framework=?", (fw,)).fetchone()[0]
            covered = conn.execute("""
                SELECT COUNT(DISTINCT s.id) FROM standards s
                JOIN cpm_standard_alignments a ON s.id = a.standard_id
                JOIN cpm_modules m ON a.module_id = m.id
                WHERE s.framework=? AND m.course_id=?
            """, (fw, course_id)).fetchone()[0]
            gaps = total_stds - covered

            cell = ws.cell(row=row, column=col, value=f"{covered}/{total_stds}")
            if covered > 0:
                cell.fill = GREEN
            col += 1

        for fw in frameworks:
            total_stds = conn.execute("SELECT COUNT(*) FROM standards WHERE framework=?", (fw,)).fetchone()[0]
            covered = conn.execute("""
                SELECT COUNT(DISTINCT s.id) FROM standards s
                JOIN cpm_standard_alignments a ON s.id = a.standard_id
                JOIN cpm_modules m ON a.module_id = m.id
                WHERE s.framework=? AND m.course_id=?
            """, (fw, course_id)).fetchone()[0]
            gaps = total_stds - covered
            cell = ws.cell(row=row, column=col, value=gaps)
            if gaps > 0:
                cell.fill = RED
            else:
                cell.fill = GREEN
            col += 1

    auto_width(ws)

    # ── Sheet 2: Coverage Matrix (MN-2022) ──
    ws2 = wb.create_sheet("Coverage Matrix - MN 2022")
    ws2.sheet_properties.tabColor = "16A34A"

    mn_stds = conn.execute(
        "SELECT * FROM standards WHERE framework='MN-2022' ORDER BY grade, code"
    ).fetchall()

    row = 1
    headers = ['MN 2022 Code', 'Grade', 'Domain', 'Description'] + courses + ['Total Hits', 'Status']
    for i, h in enumerate(headers, 1):
        ws2.cell(row=row, column=i, value=h)
    style_header_row(ws2, row, len(headers))

    for std in mn_stds:
        row += 1
        ws2.cell(row=row, column=1, value=std['code']).font = MONO
        ws2.cell(row=row, column=2, value=std['grade'])
        ws2.cell(row=row, column=3, value=std['domain'])
        desc = std['description'] or ''
        ws2.cell(row=row, column=4, value=desc[:200]).font = SMALL

        total_hits = 0
        for ci, course_id in enumerate(courses):
            hit = conn.execute("""
                SELECT COUNT(*) FROM cpm_standard_alignments a
                JOIN cpm_modules m ON a.module_id = m.id
                WHERE a.standard_id=? AND m.course_id=?
            """, (std['id'], course_id)).fetchone()[0]
            cell = ws2.cell(row=row, column=5 + ci, value=hit if hit > 0 else '')
            if hit > 0:
                cell.fill = GREEN
                cell.font = Font(bold=True, color='16A34A')
            total_hits += hit

        # Total and status
        ws2.cell(row=row, column=5 + len(courses), value=total_hits)
        status_cell = ws2.cell(row=row, column=6 + len(courses))
        if total_hits == 0:
            status_cell.value = 'GAP'
            status_cell.fill = RED
            status_cell.font = Font(bold=True, color='DC2626')
        elif total_hits > 2:
            status_cell.value = f'OVER ({total_hits}x)'
            status_cell.fill = YELLOW
            status_cell.font = Font(bold=True, color='CA8A04')
        else:
            status_cell.value = 'OK'
            status_cell.fill = GREEN
            status_cell.font = Font(bold=True, color='16A34A')

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row}"
    ws2.freeze_panes = 'A2'
    auto_width(ws2)

    # ── Sheet 3: Gap Report ──
    ws3 = wb.create_sheet("Gap Report")
    ws3.sheet_properties.tabColor = "DC2626"

    row = 1
    headers = ['Framework', 'Code', 'Grade', 'Domain', 'Description', 'DOK', 'Topic']
    for i, h in enumerate(headers, 1):
        ws3.cell(row=row, column=i, value=h)
    style_header_row(ws3, row, len(headers))

    for fw in ['MN-2022', 'CCSS-M']:
        gaps = conn.execute("""
            SELECT s.* FROM standards s
            WHERE s.framework=?
            AND s.id NOT IN (SELECT DISTINCT standard_id FROM cpm_standard_alignments)
            ORDER BY s.grade, s.code
        """, (fw,)).fetchall()

        for g in gaps:
            row += 1
            ws3.cell(row=row, column=1, value=fw)
            ws3.cell(row=row, column=2, value=g['code']).font = MONO
            ws3.cell(row=row, column=3, value=g['grade'])
            ws3.cell(row=row, column=4, value=g['domain'])
            ws3.cell(row=row, column=5, value=(g['description'] or '')[:200]).font = SMALL
            ws3.cell(row=row, column=6, value='')
            ws3.cell(row=row, column=7, value=g['topic'])

            for c in range(1, 8):
                ws3.cell(row=row, column=c).fill = RED

    ws3.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row}"
    ws3.freeze_panes = 'A2'
    auto_width(ws3)

    # ── Sheet 4: Module Checklist ──
    ws4 = wb.create_sheet("Module Checklist")
    ws4.sheet_properties.tabColor = "2563EB"

    row = 1
    headers = ['Taught?', 'Course', 'Chapter', 'Section', 'Lesson', 'Core Concepts', '# Standards', 'Standards']
    for i, h in enumerate(headers, 1):
        ws4.cell(row=row, column=i, value=h)
    style_header_row(ws4, row, len(headers))

    modules = conn.execute(
        "SELECT * FROM cpm_modules ORDER BY course_id, chapter, section, lesson"
    ).fetchall()

    for m in modules:
        row += 1
        ws4.cell(row=row, column=1, value='')  # Checkbox placeholder
        ws4.cell(row=row, column=1).fill = BLUE_LIGHT
        ws4.cell(row=row, column=2, value=m['course_id']).font = BOLD
        ws4.cell(row=row, column=3, value=m['chapter'])
        ws4.cell(row=row, column=4, value=m['section'])
        ws4.cell(row=row, column=5, value=m['lesson']).font = MONO
        ws4.cell(row=row, column=6, value=(m['core_concepts'] or '')[:100]).font = SMALL

        stds = conn.execute("""
            SELECT s.framework, s.code FROM cpm_standard_alignments a
            JOIN standards s ON a.standard_id = s.id
            WHERE a.module_id=?
        """, (m['id'],)).fetchall()

        ws4.cell(row=row, column=7, value=len(stds))
        std_str = ', '.join(f"{s['code']} ({s['framework']})" for s in stds[:10])
        if len(stds) > 10:
            std_str += f' +{len(stds)-10} more'
        ws4.cell(row=row, column=8, value=std_str).font = SMALL

    ws4.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row}"
    ws4.freeze_panes = 'A2'
    auto_width(ws4)

    # ── Per-Course Sheets ──
    for course_id in courses:
        ws_c = wb.create_sheet(f"{course_id} Detail")

        row = 1
        headers = ['Section/Lesson', 'Core Concepts', 'MN-2022 Standards', 'CCSS-M Standards', 'Core Problems', 'Notes']
        for i, h in enumerate(headers, 1):
            ws_c.cell(row=row, column=i, value=h)
        style_header_row(ws_c, row, len(headers))

        mods = conn.execute(
            "SELECT * FROM cpm_modules WHERE course_id=? ORDER BY chapter, section, lesson",
            (course_id,)
        ).fetchall()

        current_chapter = None
        for m in mods:
            if m['chapter'] != current_chapter:
                current_chapter = m['chapter']
                row += 1
                ws_c.cell(row=row, column=1, value=f"Chapter {current_chapter}").font = Font(bold=True, size=12)
                ws_c.cell(row=row, column=1).fill = GRAY

            row += 1
            ws_c.cell(row=row, column=1, value=m['lesson'] or m['section']).font = MONO
            ws_c.cell(row=row, column=2, value=m['core_concepts'] or '').font = SMALL

            mn_stds = conn.execute("""
                SELECT s.code FROM cpm_standard_alignments a
                JOIN standards s ON a.standard_id = s.id
                WHERE a.module_id=? AND s.framework='MN-2022'
            """, (m['id'],)).fetchall()
            ws_c.cell(row=row, column=3, value=', '.join(s['code'] for s in mn_stds)).font = MONO

            ccss_stds = conn.execute("""
                SELECT s.code FROM cpm_standard_alignments a
                JOIN standards s ON a.standard_id = s.id
                WHERE a.module_id=? AND s.framework='CCSS-M'
            """, (m['id'],)).fetchall()
            ws_c.cell(row=row, column=4, value=', '.join(s['code'] for s in ccss_stds)).font = MONO

            ws_c.cell(row=row, column=5, value=(m['core_problems'] or '')[:100]).font = SMALL
            ws_c.cell(row=row, column=6, value=(m['notes'] or '')[:100]).font = SMALL

            if mn_stds:
                for c in range(1, 7):
                    ws_c.cell(row=row, column=c).fill = GREEN

        auto_width(ws_c)

    # Save
    wb.save(OUT_PATH)
    conn.close()
    print(f"Exported to {OUT_PATH}")
    print(f"Sheets: {wb.sheetnames}")

if __name__ == '__main__':
    export()
