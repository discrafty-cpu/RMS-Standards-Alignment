#!/usr/bin/env python3
"""Build downloadable spreadsheets for the viewer."""

import sqlite3
import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE, "db", "curriculum.db")
OUT_DIR = os.path.join(BASE, "viewer", "downloads")

GREEN = PatternFill('solid', fgColor='DCFCE7')
RED = PatternFill('solid', fgColor='FEE2E2')
YELLOW = PatternFill('solid', fgColor='FEF9C3')
BLUE = PatternFill('solid', fgColor='DBEAFE')
GRAY = PatternFill('solid', fgColor='F3F4F6')
HDR_FILL = PatternFill('solid', fgColor='1E3A8A')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11, name='Arial')
BOLD = Font(bold=True, size=11, name='Arial')
NORMAL = Font(size=11, name='Arial')
SMALL = Font(size=10, name='Arial')
MONO = Font(name='Consolas', size=10, bold=True)
THIN = Border(
    left=Side('thin', color='DEE2E6'), right=Side('thin', color='DEE2E6'),
    top=Side('thin', color='DEE2E6'), bottom=Side('thin', color='DEE2E6'))

def hdr(ws, row, cols):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN

def auto_w(ws, max_w=55):
    for col in ws.columns:
        ml = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, max_w)

def build_all():
    os.makedirs(OUT_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    # ── 1. Master alignment workbook ──
    build_master(conn)

    # ── 2. Per-grade MN-2022 sheets ──
    for grade in ['6', '7', '8', '9']:
        build_grade_sheet(conn, grade)

    # ── 3. Per-course sheets ──
    courses = [r['id'] for r in conn.execute("SELECT id FROM cpm_courses ORDER BY id")]
    for cid in courses:
        build_course_sheet(conn, cid)

    # ── 4. Build download manifest for viewer ──
    build_manifest()

    conn.close()
    print(f"All spreadsheets saved to {OUT_DIR}")

def build_master(conn):
    wb = Workbook()

    # ── Summary sheet ──
    ws = wb.active
    ws.title = "Summary"
    ws['A1'] = "CPM Curriculum Alignment to Minnesota Standards"
    ws['A1'].font = Font(bold=True, size=16, color='1E3A8A', name='Arial')
    ws['A2'] = "Drummond Math Solutions"
    ws['A2'].font = Font(size=12, color='6B7280', name='Arial')
    ws['A3'] = "Data: CPM MN 2022 Correlations + CCSS Alignment PDFs + MN Dept of Ed Official Standards"
    ws['A3'].font = SMALL

    row = 5
    hdr(ws, row, ['Grade', 'MN-2022 Standards', 'Covered by CPM', 'Gaps', 'Coverage %', 'CPM Book(s)'])
    grade_books = {'6': 'CC1', '7': 'CC2', '8': 'CC3', '9': 'CCA, CCG, CCA2'}
    for g in ['6', '7', '8', '9']:
        row += 1
        t = conn.execute("SELECT COUNT(*) FROM standards WHERE framework='MN-2022' AND grade=?", (g,)).fetchone()[0]
        c = conn.execute("SELECT COUNT(DISTINCT s.id) FROM standards s JOIN cpm_standard_alignments a ON s.id=a.standard_id WHERE s.framework='MN-2022' AND s.grade=?", (g,)).fetchone()[0]
        ws.cell(row=row, column=1, value='Grade ' + ('9-11' if g == '9' else g)).font = BOLD
        ws.cell(row=row, column=2, value=t).font = NORMAL
        ws.cell(row=row, column=3, value=c).font = NORMAL
        ws.cell(row=row, column=3).fill = GREEN
        ws.cell(row=row, column=4, value=t - c).font = NORMAL
        ws.cell(row=row, column=4).fill = RED if t - c > 0 else GREEN
        pct = round(c / t * 100) if t else 0
        ws.cell(row=row, column=5, value=f"{pct}%").font = BOLD
        ws.cell(row=row, column=6, value=grade_books.get(g, '')).font = NORMAL
    auto_w(ws)

    # ── Per-grade sheets ──
    for g in ['6', '7', '8', '9']:
        label = f"Grade {'9-11' if g == '9' else g} MN-2022"
        ws_g = wb.create_sheet(label)
        ws_g.sheet_properties.tabColor = "1E40AF"

        row = 1
        hdr(ws_g, row, ['MN-2022 Code', 'Strand', 'Benchmark Description', 'Status',
                         'CPM Book', 'CPM Lessons', 'Alignment Source'])

        stds = conn.execute("""
            SELECT s.id, s.code, s.domain, s.description FROM standards s
            WHERE s.framework='MN-2022' AND s.grade=? ORDER BY s.code
        """, (g,)).fetchall()

        for s in stds:
            row += 1
            mods = conn.execute("""
                SELECT m.course_id, m.lesson, m.section, m.chapter, a.source
                FROM cpm_standard_alignments a
                JOIN cpm_modules m ON a.module_id = m.id
                WHERE a.standard_id=?
            """, (s['id'],)).fetchall()

            ws_g.cell(row=row, column=1, value=s['code']).font = MONO
            ws_g.cell(row=row, column=2, value=s['domain']).font = NORMAL
            ws_g.cell(row=row, column=3, value=(s['description'] or '')[:250]).font = SMALL
            ws_g.cell(row=row, column=3).alignment = Alignment(wrap_text=True)

            if mods:
                ws_g.cell(row=row, column=4, value='COVERED').font = Font(bold=True, color='15803D', name='Arial')
                ws_g.cell(row=row, column=4).fill = GREEN
                courses = sorted(set(m['course_id'] for m in mods))
                ws_g.cell(row=row, column=5, value=', '.join(courses)).font = NORMAL
                lessons = sorted(set(
                    f"{m['course_id']} {m['lesson'] or m['section'] or 'Ch' + str(m['chapter'])}"
                    for m in mods
                ))
                ws_g.cell(row=row, column=6, value=', '.join(lessons)).font = SMALL
                sources = sorted(set(m['source'] for m in mods))
                ws_g.cell(row=row, column=7, value=', '.join(sources)).font = SMALL
            else:
                ws_g.cell(row=row, column=4, value='GAP').font = Font(bold=True, color='B91C1C', name='Arial')
                ws_g.cell(row=row, column=4).fill = RED
                ws_g.cell(row=row, column=5).font = NORMAL
                ws_g.cell(row=row, column=6, value='No CPM lesson mapped').font = Font(italic=True, color='6B7280', name='Arial')

            for c in range(1, 8):
                ws_g.cell(row=row, column=c).border = THIN

        ws_g.auto_filter.ref = f"A1:G{row}"
        ws_g.freeze_panes = 'A2'
        auto_w(ws_g)
        ws_g.column_dimensions['C'].width = 55

    # ── Per-course sheets ──
    courses = conn.execute("SELECT * FROM cpm_courses WHERE id IN ('CC1','CC2','CC3','CCA','CCG','CCA2') ORDER BY id").fetchall()
    for course in courses:
        cid = course['id']
        ws_c = wb.create_sheet(f"{cid} Alignment")
        ws_c.sheet_properties.tabColor = "7C3AED"

        row = 1
        hdr(ws_c, row, ['Chapter', 'Lesson/Section', 'Core Concepts',
                         'MN-2022 Standards', 'CCSS-M Standards', 'Alignment Source'])

        mods = conn.execute(
            "SELECT * FROM cpm_modules WHERE course_id=? ORDER BY chapter, section, lesson",
            (cid,)).fetchall()

        cur_ch = None
        for m in mods:
            if m['chapter'] != cur_ch:
                cur_ch = m['chapter']
                row += 1
                ws_c.cell(row=row, column=1, value=f"Chapter {cur_ch}").font = BOLD
                for c in range(1, 7):
                    ws_c.cell(row=row, column=c).fill = PatternFill('solid', fgColor='1E3A8A')
                    ws_c.cell(row=row, column=c).font = Font(bold=True, color='FFFFFF', name='Arial')

            row += 1
            ws_c.cell(row=row, column=1, value=m['chapter']).font = NORMAL
            ws_c.cell(row=row, column=2, value=m['lesson'] or m['section'] or '').font = MONO
            ws_c.cell(row=row, column=3, value=(m['core_concepts'] or '')[:100]).font = SMALL

            mn_stds = conn.execute("""
                SELECT s.code, a.source FROM cpm_standard_alignments a
                JOIN standards s ON a.standard_id = s.id
                WHERE a.module_id=? AND s.framework='MN-2022'
            """, (m['id'],)).fetchall()
            ccss_stds = conn.execute("""
                SELECT s.code, a.source FROM cpm_standard_alignments a
                JOIN standards s ON a.standard_id = s.id
                WHERE a.module_id=? AND s.framework='CCSS-M'
            """, (m['id'],)).fetchall()

            mn_str = ', '.join(sorted(set(s['code'] for s in mn_stds)))
            ccss_str = ', '.join(sorted(set(s['code'] for s in ccss_stds)))
            sources = sorted(set(s['source'] for s in mn_stds + ccss_stds))

            ws_c.cell(row=row, column=4, value=mn_str).font = SMALL
            ws_c.cell(row=row, column=5, value=ccss_str).font = SMALL
            ws_c.cell(row=row, column=6, value=', '.join(sources)).font = SMALL

            if mn_str:
                ws_c.cell(row=row, column=4).fill = GREEN
            if ccss_str:
                ws_c.cell(row=row, column=5).fill = BLUE

            for c in range(1, 7):
                ws_c.cell(row=row, column=c).border = THIN

        ws_c.auto_filter.ref = f"A1:F{row}"
        ws_c.freeze_panes = 'A2'
        auto_w(ws_c)

    path = os.path.join(OUT_DIR, "CPM_Standards_Alignment_Master.xlsx")
    wb.save(path)
    print(f"  Master: {path} ({len(wb.sheetnames)} sheets)")

def build_grade_sheet(conn, grade):
    wb = Workbook()
    ws = wb.active
    label = '9-11' if grade == '9' else grade
    ws.title = f"Grade {label} MN-2022"

    row = 1
    ws.cell(row=1, column=1, value=f"Grade {label} MN 2022 Standards — CPM Alignment").font = Font(bold=True, size=14, name='Arial')
    row = 3
    hdr(ws, row, ['Code', 'Strand', 'Benchmark', 'Status', 'CPM Lessons', 'Source'])

    for s in conn.execute("SELECT * FROM standards WHERE framework='MN-2022' AND grade=? ORDER BY code", (grade,)):
        row += 1
        mods = conn.execute("""
            SELECT m.course_id, m.lesson, m.section, m.chapter, m.core_concepts, a.source
            FROM cpm_standard_alignments a JOIN cpm_modules m ON a.module_id = m.id
            WHERE a.standard_id=?
        """, (s['id'],)).fetchall()

        ws.cell(row=row, column=1, value=s['code']).font = MONO
        ws.cell(row=row, column=2, value=s['domain']).font = NORMAL
        ws.cell(row=row, column=3, value=(s['description'] or '')).font = SMALL
        ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)

        if mods:
            ws.cell(row=row, column=4, value='COVERED').fill = GREEN
            ws.cell(row=row, column=4).font = Font(bold=True, color='15803D', name='Arial')
            lessons = [f"{m['course_id']} {m['lesson'] or m['section'] or 'Ch'+str(m['chapter'])} ({(m['core_concepts'] or '')[:30]})"
                       for m in mods]
            ws.cell(row=row, column=5, value='; '.join(sorted(set(lessons)))).font = SMALL
            ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=6, value=', '.join(sorted(set(m['source'] for m in mods)))).font = SMALL
        else:
            ws.cell(row=row, column=4, value='GAP').fill = RED
            ws.cell(row=row, column=4).font = Font(bold=True, color='B91C1C', name='Arial')

        for c in range(1, 7):
            ws.cell(row=row, column=c).border = THIN

    ws.auto_filter.ref = f"A3:F{row}"
    ws.freeze_panes = 'A4'
    auto_w(ws)
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['E'].width = 50

    path = os.path.join(OUT_DIR, f"Grade_{label}_MN2022_Alignment.xlsx")
    wb.save(path)
    print(f"  Grade {label}: {path}")

def build_course_sheet(conn, cid):
    course = conn.execute("SELECT * FROM cpm_courses WHERE id=?", (cid,)).fetchone()
    if not course:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = f"{cid} Alignment"

    ws.cell(row=1, column=1, value=f"{course['name']} ({cid}) — Standards Alignment").font = Font(bold=True, size=14, name='Arial')

    row = 3
    hdr(ws, row, ['Ch', 'Lesson', 'Core Concepts', 'MN-2022', 'CCSS-M', 'All Standards', 'Source'])

    mods = conn.execute("SELECT * FROM cpm_modules WHERE course_id=? ORDER BY chapter, section, lesson", (cid,)).fetchall()
    cur_ch = None
    for m in mods:
        if m['chapter'] != cur_ch:
            cur_ch = m['chapter']
            row += 1
            ws.cell(row=row, column=1, value=f"Chapter {cur_ch}").font = BOLD
            for c in range(1, 8):
                ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor='1E3A8A')
                ws.cell(row=row, column=c).font = Font(bold=True, color='FFFFFF', name='Arial')

        row += 1
        ws.cell(row=row, column=1, value=m['chapter']).font = NORMAL
        ws.cell(row=row, column=2, value=m['lesson'] or m['section'] or '').font = MONO
        ws.cell(row=row, column=3, value=(m['core_concepts'] or '')).font = SMALL
        ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)

        all_stds = conn.execute("""
            SELECT s.framework, s.code, a.source FROM cpm_standard_alignments a
            JOIN standards s ON a.standard_id = s.id WHERE a.module_id=?
        """, (m['id'],)).fetchall()

        mn = sorted(set(s['code'] for s in all_stds if s['framework'] == 'MN-2022'))
        ccss = sorted(set(s['code'] for s in all_stds if s['framework'] == 'CCSS-M'))
        all_codes = sorted(set(f"{s['code']} ({s['framework']})" for s in all_stds))
        sources = sorted(set(s['source'] for s in all_stds))

        ws.cell(row=row, column=4, value=', '.join(mn)).font = SMALL
        if mn: ws.cell(row=row, column=4).fill = GREEN
        ws.cell(row=row, column=5, value=', '.join(ccss)).font = SMALL
        if ccss: ws.cell(row=row, column=5).fill = BLUE
        ws.cell(row=row, column=6, value=', '.join(all_codes)).font = SMALL
        ws.cell(row=row, column=7, value=', '.join(sources)).font = SMALL

        for c in range(1, 8):
            ws.cell(row=row, column=c).border = THIN

    ws.auto_filter.ref = f"A3:G{row}"
    ws.freeze_panes = 'A4'
    auto_w(ws)
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['F'].width = 50

    path = os.path.join(OUT_DIR, f"{cid}_Alignment.xlsx")
    wb.save(path)
    print(f"  {cid}: {path}")

def build_manifest():
    files = []
    for f in sorted(os.listdir(OUT_DIR)):
        if f.endswith('.xlsx'):
            size = os.path.getsize(os.path.join(OUT_DIR, f))
            files.append({'filename': f, 'size': size})
    manifest = os.path.join(OUT_DIR, 'manifest.json')
    with open(manifest, 'w') as fh:
        json.dump(files, fh, indent=2)
    print(f"  Manifest: {len(files)} files")

if __name__ == '__main__':
    build_all()
